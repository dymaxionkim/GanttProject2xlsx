import pandas
import pandas as pd 
from datetime import datetime, timedelta

# gantt
gantt = pandas.read_csv('gantt.csv')
gantt.rename(columns={'완료':'진행률','코디네이터':'담당자', '개요 번호':'개요번호', '선행 작업':'선행작업', '작업 색상':'작업색상'}, inplace=True)

# Size
gantt_index = len(gantt.index)
gantt_column = len(gantt.columns)

# gantt2
gantt2 = pd.DataFrame(index=range(gantt_index), columns=['번호','목표업무','담당자','시작일','종료일','진행일수','남은날짜','노트','진행률'])
gantt2.번호 = gantt.개요번호
gantt2.목표업무 = gantt.이름
gantt2.담당자 = gantt.담당자
gantt2.시작일 = gantt.시작일
gantt2.종료일 = gantt.종료일
gantt2.노트 = gantt.노트
gantt2.진행률 = gantt.진행률

#gantt2.진행일수,남은날짜
TODAY = datetime.today()
for i in range(gantt_index):
    STARTDAY = datetime.strptime(gantt2.시작일[i],'%Y-%m-%d')
    ENDDAY = datetime.strptime(gantt2.종료일[i],'%Y-%m-%d')
    gantt2.진행일수[i] = (ENDDAY-STARTDAY).days + 1
    gantt2.남은날짜[i] = (ENDDAY-TODAY).days
    
# gantt3
WEEKTODAY = datetime.today().weekday()
gantt3 = pd.DataFrame(index=range(gantt_index),columns=range(21))
gantt3_index = len(gantt3.index)
gantt3_column = len(gantt3.columns)
for j in range(gantt3_index):
    for i in range(gantt3_column):
        gantt3[i][j] = 0
        STARTDAY = datetime.strptime(gantt2.시작일[j],'%Y-%m-%d')
        ENDDAY = datetime.strptime(gantt2.종료일[j],'%Y-%m-%d')
        THIS_DAY = TODAY+timedelta(days=i-(WEEKTODAY+7))
        BEFORE_DAYS = int((STARTDAY-THIS_DAY).days)
        LAST_DAYS = int((ENDDAY-THIS_DAY).days)+1
        if BEFORE_DAYS<0 and LAST_DAYS>=0:
            gantt3[i][j] = 1

# gantt4
gantt4 = pd.concat([gantt2,gantt3],axis=1)

#########################################################
# openpyxl
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

# New Book
book = Workbook()
sheet = book.active
sheet.title = "일일업무보고"

# Copy DataFrame to Sheet
rows = dataframe_to_rows(gantt4)
for i, row in enumerate(rows,1):
    for j, value in enumerate(row,1):
         sheet.cell(row=i,column=j,value=value)
            
# Width
sheet.column_dimensions['A'].width = 5
sheet.column_dimensions['B'].width = 5
sheet.column_dimensions['C'].width = 40
sheet.column_dimensions['D'].width = 10
sheet.column_dimensions['E'].width = 12
sheet.column_dimensions['F'].width = 12
sheet.column_dimensions['G'].width = 8
sheet.column_dimensions['H'].width = 8
sheet.column_dimensions['I'].width = 40
sheet.column_dimensions['J'].width = 10

for temp in ['K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE']:
    sheet.column_dimensions[temp].width = 3
    
# Align
for i in list(range(1,sheet.max_column+1)):
    for j in list(range(1,sheet.max_row+1)):
        sheet.cell(column=i,row=j).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
for i in [1,2]:
    for j in list(range(1,sheet.max_row+1)):
        sheet.cell(column=i,row=j).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
for i in [3,9]:
    for j in list(range(1,sheet.max_row+1)):
        sheet.cell(column=i,row=j).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
for i in list(range(1,sheet.max_column+1)):
    for j in [2]:
        sheet.cell(column=i,row=j).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
for i in list(range(1,sheet.max_column+1)):
    for j in [1]:
        sheet.cell(column=i,row=j).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        
# Border
for i in list(range(1,sheet.max_column+1)):
    for j in list(range(2,sheet.max_row+1)):
        sheet.cell(column=i,row=j).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))  

# Colors
GANTT_COLOR = 'FFDDDDFF'
GANTT_TODAY_COLOR = 'FF7878E1'
GANTT_TODAY_FINISHED_COLOR = 'FF000000'
GANTT_DELAYED_COLOR = 'FF98A3'
WHITE_COLOR = 'FFFFFFFF'
SUBJECTS_COLOR = 'FFEEEEEE'
PROJECT_COLOR = 'FFDDDDDD'
FINISHED_FONT_COLOR = "FF999999"

# Header
WEEKNAME = ['월','화','수','목','금','토','일']
for i in list(range(1,sheet.max_column+1)):
    sheet.cell(column=i,row=1).value = ""
    sheet.cell(column=i,row=1).font = openpyxl.styles.fonts.Font(bold=True)
sheet.cell(column=2,row=1).value = datetime.today().strftime("%Y-%m-%d") + " (" + WEEKNAME[WEEKTODAY] + ") " + "기구팀_업무현황_요약보고"

# Subjects
SUBJECTS = ['','번호','목표업무','담당자','시작일','종료일','진행일수','남은날짜','노트','진행률','월','화','수','목','금','토','일','월','화','수','목','금','토','일','월','화','수','목','금','토','일']
for i in list(range(len(SUBJECTS))):
    sheet.cell(column=i+1,row=2).value = SUBJECTS[i]
    sheet.cell(column=i+1,row=2).font = openpyxl.styles.fonts.Font(bold=True)
    sheet.cell(column=i+1,row=2).fill = PatternFill(start_color=SUBJECTS_COLOR,end_color=SUBJECTS_COLOR,fill_type='solid')

# Sub-Subjects
sheet.merge_cells('K1:Q1')
sheet.merge_cells('R1:X1')
sheet.merge_cells('Y1:AE1')
SUBJECTS = ['지난주','금주','다음주']
TEMP = [11,18,25]
for i in range(3):
    sheet.cell(column=TEMP[i],row=1).value = SUBJECTS[i]
    sheet.cell(column=TEMP[i],row=1).fill = PatternFill(start_color=SUBJECTS_COLOR,end_color=SUBJECTS_COLOR,fill_type='solid')
    sheet.cell(column=TEMP[i],row=1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
for i in list(range(11,32)):
    sheet.cell(column=i,row=1).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

# Project Head
for i in list(range(sheet.max_column-21)):
    for j in list(range(2,sheet.max_row)):
        if float(sheet.cell(column=2,row=j+1).value)%1==0:
            sheet.cell(column=i+1,row=j+1).font = openpyxl.styles.fonts.Font(bold=True)
            sheet.cell(column=i+1,row=j+1).fill = PatternFill(start_color=PROJECT_COLOR,end_color=PROJECT_COLOR,fill_type='solid')
            
# Gantt
START_COLUMN = 11
START_ROW = 3
END_COLUMN = sheet.max_column+1
END_ROW = sheet.max_row+1
WEEKTODAY_COLUMN = WEEKTODAY+18
sheet.cell(column=WEEKTODAY_COLUMN,row=2).fill = PatternFill(start_color=GANTT_TODAY_COLOR,end_color=GANTT_TODAY_COLOR,fill_type='solid')
for i in list(range(START_COLUMN,END_COLUMN)):
    for j in list(range(START_ROW,END_ROW)):
        # 작업일 색상 강조
        if sheet.cell(column=i,row=j).value==1:
            sheet.cell(column=i,row=j).fill = PatternFill(start_color=GANTT_COLOR,end_color=GANTT_COLOR,fill_type='solid')
            # 금일 색상 강조
            if WEEKTODAY_COLUMN==i:
                sheet.cell(column=i,row=j).fill = PatternFill(start_color=GANTT_TODAY_COLOR,end_color=GANTT_TODAY_COLOR,fill_type='solid')
                sheet.cell(column=i,row=2).font = Font(color = WHITE_COLOR)
        # 지연된 작업 색상 강조
        if sheet.cell(column=i,row=j).value==1:
            if sheet.cell(column=10,row=j).value<100 and sheet.cell(column=8,row=j).value<-1:
                sheet.cell(column=i,row=j).fill = PatternFill(start_color=GANTT_DELAYED_COLOR,end_color=GANTT_DELAYED_COLOR,fill_type='solid')
                
# Gantt Font Size
for i in list(range(START_COLUMN,END_COLUMN)):
    for j in list(range(START_ROW,END_ROW)):
        sheet.cell(column=i,row=j).font = Font(size=8)
        
# Gantt 진행률
for i in list(range(START_COLUMN,END_COLUMN)):
    for j in list(range(START_ROW,END_ROW)):
        if sheet.cell(column=10,row=j).value==100 and sheet.cell(column=i,row=j).value==1:
            # 완료된 작업 100 표기
            if sheet.cell(column=i+1,row=j).value==0:
                sheet.cell(column=i,row=j).value=sheet.cell(column=10,row=j).value
            # 금일 완료된 작업 색상 강조
            if WEEKTODAY_COLUMN==i:
                sheet.cell(column=i,row=j).font = Font(color=WHITE_COLOR,size=8)
                sheet.cell(column=i,row=j).fill = PatternFill(start_color=GANTT_TODAY_FINISHED_COLOR,end_color=GANTT_TODAY_FINISHED_COLOR,fill_type='solid')
        # 금일 진행률 표기
        if WEEKTODAY_COLUMN==i and sheet.cell(column=i,row=j).value==1:
            sheet.cell(column=i,row=j).value=sheet.cell(column=10,row=j).value
            sheet.cell(column=i,row=j).font = Font(color=WHITE_COLOR,size=8)
        # 지연된 작업 진행률 표기
        if sheet.cell(column=10,row=j).value<100 and sheet.cell(column=8,row=j).value<-1:
            if sheet.cell(column=i+1,row=j).value==0 and sheet.cell(column=i,row=j).value==1:
                sheet.cell(column=i,row=j).value=sheet.cell(column=10,row=j).value
            
# Finished Tasks
for i in list(range(1,START_COLUMN-1)):
    for j in list(range(START_ROW,END_ROW)):
        if sheet.cell(column=10,row=j).value==100 and float(sheet.cell(column=2,row=j).value)%1!=0:
            if sheet.cell(column=8,row=j).value!=-1:
                sheet.cell(column=i,row=j).font = Font(color=FINISHED_FONT_COLOR)
            
# Gantt Remove Flag
for i in list(range(START_COLUMN,END_COLUMN)):
    for j in list(range(START_ROW,END_ROW)):
        if sheet.cell(column=i,row=j).value==0 or sheet.cell(column=i,row=j).value==1:
            sheet.cell(column=i,row=j).value=""

# Folding column
sheet.column_dimensions['A'].hidden = True
sheet.column_dimensions['E'].hidden = True
sheet.column_dimensions['G'].hidden = True
sheet.column_dimensions['J'].hidden = True

# Folding rows passing out from today
for j in list(range(START_ROW,END_ROW)):
    # 100% 완료된 작업들
    if sheet.cell(column=10,row=j).value==100: 
        sheet.row_dimensions[j].hidden = True
        # 프로젝트 제목은 살려주기
        if float(sheet.cell(column=2,row=j).value)%1==0:
            sheet.row_dimensions[j].hidden = False
        # 완료일자가 1주일 안 지난 것들은 살려주기
        if sheet.cell(column=8,row=j).value+7>0:
            sheet.row_dimensions[j].hidden = False
    # 0% 작업들
    if sheet.cell(column=10,row=j).value==0:
        sheet.row_dimensions[j].hidden = True
        # 프로젝트 제목은 살려주기
        if float(sheet.cell(column=2,row=j).value)%1==0:
            sheet.row_dimensions[j].hidden = False

# Change 번호
for j in list(range(START_ROW,END_ROW)):
    if sheet.cell(column=2,row=j).value%1!=0:
        #sheet.cell(column=2,row=j).value = str(sheet.cell(column=2,row=j).value)
        sheet.cell(column=2,row=j).value = "-"

# 범례
REMARK_CULUMN = 11
REMARK_ROW = sheet.max_row+2
REMARK_SPACE = 4

sheet.cell(column=REMARK_CULUMN-2,row=REMARK_ROW).alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
sheet.cell(column=REMARK_CULUMN-2,row=REMARK_ROW).value = "○ 범례 :  "

sheet.cell(column=REMARK_CULUMN,row=REMARK_ROW).fill = PatternFill(start_color=GANTT_TODAY_COLOR,end_color=GANTT_TODAY_COLOR,fill_type='solid')
sheet.cell(column=REMARK_CULUMN,row=REMARK_ROW).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
sheet.cell(column=REMARK_CULUMN,row=REMARK_ROW).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet.cell(column=REMARK_CULUMN,row=REMARK_ROW).font = Font(color=WHITE_COLOR,size=8)
sheet.cell(column=REMARK_CULUMN,row=REMARK_ROW).value = "00"
sheet.cell(column=REMARK_CULUMN+1,row=REMARK_ROW).value = "진행률"

sheet.cell(column=REMARK_CULUMN+REMARK_SPACE,row=REMARK_ROW).fill = PatternFill(start_color=GANTT_COLOR,end_color=GANTT_COLOR,fill_type='solid')
sheet.cell(column=REMARK_CULUMN+REMARK_SPACE,row=REMARK_ROW).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
sheet.cell(column=REMARK_CULUMN+REMARK_SPACE,row=REMARK_ROW).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet.cell(column=REMARK_CULUMN+REMARK_SPACE,row=REMARK_ROW).font = Font(size=8)
sheet.cell(column=REMARK_CULUMN+REMARK_SPACE,row=REMARK_ROW).value = "100"
sheet.cell(column=REMARK_CULUMN+REMARK_SPACE+1,row=REMARK_ROW).value = "완료시점"

sheet.cell(column=REMARK_CULUMN+2*REMARK_SPACE,row=REMARK_ROW).fill = PatternFill(start_color=GANTT_TODAY_FINISHED_COLOR,end_color=GANTT_TODAY_FINISHED_COLOR,fill_type='solid')
sheet.cell(column=REMARK_CULUMN+2*REMARK_SPACE,row=REMARK_ROW).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
sheet.cell(column=REMARK_CULUMN+2*REMARK_SPACE,row=REMARK_ROW).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet.cell(column=REMARK_CULUMN+2*REMARK_SPACE,row=REMARK_ROW).font = Font(color=WHITE_COLOR,size=8)
sheet.cell(column=REMARK_CULUMN+2*REMARK_SPACE,row=REMARK_ROW).value = "100"
sheet.cell(column=REMARK_CULUMN+2*REMARK_SPACE+1,row=REMARK_ROW).value = "금일완료"

sheet.cell(column=REMARK_CULUMN+3*REMARK_SPACE,row=REMARK_ROW).fill = PatternFill(start_color=GANTT_DELAYED_COLOR,end_color=GANTT_DELAYED_COLOR,fill_type='solid')
sheet.cell(column=REMARK_CULUMN+3*REMARK_SPACE,row=REMARK_ROW).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
sheet.cell(column=REMARK_CULUMN+3*REMARK_SPACE,row=REMARK_ROW).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet.cell(column=REMARK_CULUMN+3*REMARK_SPACE,row=REMARK_ROW).font = Font(size=8)
sheet.cell(column=REMARK_CULUMN+3*REMARK_SPACE,row=REMARK_ROW).value = "00"
sheet.cell(column=REMARK_CULUMN+3*REMARK_SPACE+1,row=REMARK_ROW).value = "지연작업"

sheet.cell(column=REMARK_CULUMN+4*REMARK_SPACE,row=REMARK_ROW).border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
sheet.cell(column=REMARK_CULUMN+4*REMARK_SPACE,row=REMARK_ROW).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet.cell(column=REMARK_CULUMN+4*REMARK_SPACE,row=REMARK_ROW).font = Font(color=FINISHED_FONT_COLOR)
sheet.cell(column=REMARK_CULUMN+4*REMARK_SPACE,row=REMARK_ROW).value = "완"
sheet.cell(column=REMARK_CULUMN+4*REMARK_SPACE+1,row=REMARK_ROW).value = "완료작업"

# Sheet Page Settup
sheet.title = '기구팀'
sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.scale = 68
sheet.page_setup.fitToWidth = 1
sheet.page_setup.fitToHeight = 0

book.save('Report.xlsx')
